import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime
import pytz
import importlib

st.set_page_config(page_title="Walmart Orders Export", layout="wide")
st.title("Walmart Orders Export")
st.markdown("Upload SPS order file in original format.")

uploaded_file = st.file_uploader("Upload Walmart File (.csv or .xlsx)", type=["csv", "xlsx"])

MAP_BC = {
    "665069485": ("B8100217","(BURPEE), (ECO FRIEND), SEED STARTING MIX, 0.06-0.03-0.03, 12 QT"),
    "665113710": ("B8100217","(BURPEE), (ECO FRIEND), SEED STARTING MIX, 0.06-0.03-0.03, 12 QT"),
    "666192291": ("B8100925","(BURPEE), (ECO FRIEND), SEED STARTING MIX, 0.06-0.03-0.03, 16 QT"),
    "665069486": ("B8100925","(BURPEE), (ECO FRIEND), SEED STARTING MIX, 0.06-0.03-0.03, 16 QT"),
    "671533940": ("B8100204","(BURPEE), (NAT ORG), ECO FRIENDLY SEED START MIX COIR - BRICK, 8 QT"),
    "665029761": ("B8100914","3-3-2 (EXPERT GARDENER ORG), CHICKEN MANURE, PF, 4 LB."),
    "665031601": ("B8100914","3-3-2 (EXPERT GARDENER ORG), CHICKEN MANURE, PF, 4 LB."),
    "665029760": ("B8100930","3-5-6 (EXPERT GARDENER ORG), VEG TOMATO, PF, 4 LB."),
    "665031679": ("B8100930","3-5-6 (EXPERT GARDENER ORG), VEG TOMATO, PF, 4 LB."),
    "665031685": ("B8100929","3-5-6 (EXPERT GARDENER ORG), VEG TOMATO, PF, 8 LB"),
    "665029764": ("B8100929","3-5-6 (EXPERT GARDENER ORG), VEG TOMATO, PF, 8 LB"),
    "665031697": ("B8100932","4-4-4 (EXPERT GARDENER ORG), ALL PURPOSE, PF, 4 LB."),
    "665029763": ("B8100932","4-4-4 (EXPERT GARDENER ORG), ALL PURPOSE, PF, 4 LB."),
    "665031676": ("B8100928","4-4-4 (EXPERT GARDENER ORG), ALL PURPOSE, PF, 8 LB."),
    "665029762": ("B8100928","4-4-4 (EXPERT GARDENER ORG), ALL PURPOSE, PF, 8 LB."),
    "665106936": ("B1258730","SUNN 6-6-6 33#"),
    "656676067": ("B1201480","SUNN BLOOM 2-10-10 20#"),
    "565380341": ("B1201480","SUNN BLOOM 2-10-10 20#"),
    "565378804": ("B1202370","SUNN CITRUS 6-4-6 4/10#"),
    "656676244": ("B1202380","SUNN CITRUS FERTILIZER 20#"),
    "565380344": ("B1202380","SUNN CITRUS FERTILIZER 20#"),
    "656676073": ("B1224080","SUNN GARD/AZA/CAM 8-4-8 20#"),
    "565380342": ("B1224080","SUNN GARD/AZA/CAM 8-4-8 20#"),
    "565378805": ("B1224070","SUNN GARD/AZA/CAM 8-4-8 4/10#"),
    "665106990": ("B1251580","SUNN NITRO GREEN 16-0-8 33#"),
    "656679467": ("B1260080","SUNN PALM 6-1-8 20 LB"),
    "565380343": ("B1260080","SUNN PALM 6-1-8 20 LB"),
    "565378806": ("B1260070","SUNN PALM 6-1-8 4/10#"),
}

OUTPUT_COLUMNS = [
    "PO Number","PO Date","Retailers PO","Ship Dates","Cancel Date","PO Line #",
    "BC Item#","BC Item Name","Qty Ordered","Unit of Measure","Unit Price",
    "Buyers Catalog or Stock Keeping #","UPC/EAN","Vendor Style","Number of Inner Packs",
    "Vendor #","Promo #","Ticket Description","Other Info / #s","Frt Terms","Payment Terms %",
    "Payment Terms Disc Days Due","Payment Terms Net Days","Allow/Charge Type","Allow/Charge Service",
    "Allow/Charge Amt","Allow/Charge %","Buying Party Name","Buying Party Location",
    "Buying Party Address 1","Buying Party Address 2","Buying Party City","Buying Party State",
    "Buying Party Zip","Buying Party Country","Notes/Comments","GTIN","PO Total Amount",
    "Must Arrive By","EDITxnType"
]

DATE_COLS = ["PO Date","Ship Dates","Cancel Date","Must Arrive By"]
INT_COLS = ["Qty Ordered","Number of Inner Packs","PO Line #","Payment Terms Disc Days Due","Payment Terms Net Days"]
FLOAT_COLS = ["Unit Price","Allow/Charge Amt","Allow/Charge %","Payment Terms %","PO Total Amount"]

def can_import(mod: str) -> bool:
    return importlib.util.find_spec(mod) is not None

def coalesce_cols(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [c.strip() for c in df.columns]
    for c in df.columns:
        s = df[c].astype(str).str.strip()
        df[c] = s.mask(s.isin(["", "nan", "None"]))
    return df

def apply_bc_mapping(df: pd.DataFrame) -> pd.DataFrame:
    keycol = "Buyers Catalog or Stock Keeping #"
    if keycol in df.columns:
        key = df[keycol].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
        df["BC Item#"] = key.map(lambda k: MAP_BC.get(k, (None, None))[0])
        df["BC Item Name"] = key.map(lambda k: MAP_BC.get(k, (None, None))[1])
    else:
        df["BC Item#"] = None
        df["BC Item Name"] = None
    return df

def to_datetime_cols(df: pd.DataFrame, cols) -> pd.DataFrame:
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce")
    return df

def to_numeric_cols(df: pd.DataFrame, cols, is_int=False) -> pd.DataFrame:
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
            if is_int:
                df[c] = df[c].dropna().astype("Int64").reindex(df.index)
    return df

def finalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    for c in OUTPUT_COLUMNS:
        if c not in df.columns:
            df[c] = pd.NA
    return df[OUTPUT_COLUMNS]

if uploaded_file:
    name = uploaded_file.name.lower()
    if name.endswith(".xlsx") and not can_import("openpyxl"):
        st.error("XLSX reading requires openpyxl. Add it to requirements.txt or upload a CSV.")
        st.stop()

    try:
        if name.endswith(".csv"):
            df = pd.read_csv(uploaded_file, dtype=str)
        else:
            df = pd.read_excel(uploaded_file, dtype=str)
    except Exception as e:
        st.error(f"Could not read file: {e}")
        st.stop()

    df = coalesce_cols(df)
    df = apply_bc_mapping(df)
    df = to_datetime_cols(df, DATE_COLS)              # Excel-recognizable dates (esp. PO Date)
    df = to_numeric_cols(df, INT_COLS, is_int=True)   # integers
    df = to_numeric_cols(df, FLOAT_COLS, is_int=False)# numeric amounts/percents as numbers
    df = finalize_columns(df)

    tz = pytz.timezone("America/New_York")
    timestamp = datetime.now(tz).strftime("%m.%d.%Y_%H.%M")
    base = f"Walmart_Export_{timestamp}"

    can_xlsxwriter = can_import("xlsxwriter")
    can_openpyxl = can_import("openpyxl")
    output = BytesIO()

    if can_xlsxwriter or can_openpyxl:
        engine = "xlsxwriter" if can_xlsxwriter else "openpyxl"
        try:
            with pd.ExcelWriter(output, engine=engine) as writer:
                df.to_excel(writer, index=False, sheet_name="Export")

                if engine == "xlsxwriter":
                    wb = writer.book
                    ws = writer.sheets["Export"]
                    fmt_text = wb.add_format({"font_name":"Aptos Narrow","font_size":11,"align":"left"})
                    fmt_header = wb.add_format({"font_name":"Aptos Narrow","font_size":11,"bold":True,"align":"left"})
                    fmt_date = wb.add_format({"font_name":"Aptos Narrow","font_size":11,"align":"left","num_format":"mm/dd/yyyy"})
                    fmt_int = wb.add_format({"font_name":"Aptos Narrow","font_size":11,"align":"left","num_format":"0"})
                    fmt_float = wb.add_format({"font_name":"Aptos Narrow","font_size":11,"align":"left","num_format":"0.00"})

                    for col_idx, c in enumerate(OUTPUT_COLUMNS):
                        ws.set_column(col_idx, col_idx, 22, fmt_text)
                        ws.write(0, col_idx, c, fmt_header)

                    for c in DATE_COLS:
                        if c in OUTPUT_COLUMNS:
                            idx = OUTPUT_COLUMNS.index(c)
                            ws.set_column(idx, idx, 16, fmt_date)

                    for c in INT_COLS:
                        if c in OUTPUT_COLUMNS:
                            idx = OUTPUT_COLUMNS.index(c)
                            ws.set_column(idx, idx, 14, fmt_int)

                    for c in FLOAT_COLS:
                        if c in OUTPUT_COLUMNS:
                            idx = OUTPUT_COLUMNS.index(c)
                            ws.set_column(idx, idx, 16, fmt_float)

                else:
                    from openpyxl.styles import Font, Alignment, NamedStyle
                    from openpyxl.utils import get_column_letter
                    wb = writer.book
                    ws = writer.sheets["Export"]

                    font = Font(name="Aptos Narrow", size=11, bold=False)
                    font_bold = Font(name="Aptos Narrow", size=11, bold=True)
                    left = Alignment(horizontal="left")

                    for j, col in enumerate(OUTPUT_COLUMNS, start=1):
                        cell = ws.cell(row=1, column=j)
                        cell.font = font_bold
                        cell.alignment = left
                        ws.column_dimensions[get_column_letter(j)].width = 22

                    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        for cell in r:
                            cell.font = font
                            cell.alignment = left

                    for c in DATE_COLS:
                        if c in OUTPUT_COLUMNS:
                            j = OUTPUT_COLUMNS.index(c) + 1
                            for i in range(2, ws.max_row + 1):
                                ws.cell(row=i, column=j).number_format = "mm/dd/yyyy"

                    for c in INT_COLS:
                        if c in OUTPUT_COLUMNS:
                            j = OUTPUT_COLUMNS.index(c) + 1
                            for i in range(2, ws.max_row + 1):
                                ws.cell(row=i, column=j).number_format = "0"

                    for c in FLOAT_COLS:
                        if c in OUTPUT_COLUMNS:
                            j = OUTPUT_COLUMNS.index(c) + 1
                            for i in range(2, ws.max_row + 1):
                                ws.cell(row=i, column=j).number_format = "0.00"

            st.success("File processed successfully.")
            st.download_button(
                "Download Walmart Export (Excel)",
                data=output.getvalue(),
                file_name=base + ".xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            csv_bytes = df.to_csv(index=False).encode("utf-8-sig")
            st.warning(f"Excel writer unavailable ({e}). Providing CSV instead.")
            st.download_button(
                "Download Walmart Export (CSV)",
                data=csv_bytes,
                file_name=base + ".csv",
                mime="text/csv",
            )
    else:
        csv_bytes = df.to_csv(index=False).encode("utf-8-sig")
        st.info("Excel engines not installed. Providing CSV download.")
        st.download_button(
            "Download Walmart Export (CSV)",
            data=csv_bytes,
            file_name=base + ".csv",
            mime="text/csv",
        )
